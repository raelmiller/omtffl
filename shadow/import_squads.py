#!/usr/bin/env python3
"""Turn the auction app's Excel export into squads.json.

Resolves each drafted player to their FPL element id, which is what the
scoring engine works in. The auction app records FPL's own `web_name`, so
these match exactly rather than needing the fuzzy matching the older draft
workbooks required — but anything that doesn't resolve is reported loudly
rather than quietly dropped, because a silently missing player would just
show up later as a team mysteriously scoring less.

Usage:
    python3 shadow/import_squads.py path/to/fpl-draft-results.xlsx
"""
import json
import sys
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parent
FEED = ROOT.parent / "players-feed" / "players.json"
OUT = ROOT / "data" / "squads.json"
IDENTITIES = ROOT / "data" / "identities.json"

POS = {1: "GK", 2: "DEF", 3: "MID", 4: "FWD"}


def load_identities():
    """Map the draft export's owner names to public-safe identities.

    The export records whoever ran the draft typed in — first names. This maps
    them to team name plus initials so nothing downstream carries a person's
    name, which is what lets the squads live in a public repo.
    """
    if not IDENTITIES.exists():
        sys.exit(
            f"No {IDENTITIES.name} — it maps draft owner names to team/initials.\n"
            'Format: {"identities": {"<owner>": {"key": "AB", "team": "Team Name"}}}'
        )
    return json.loads(IDENTITIES.read_text())["identities"]


def load_feed():
    if not FEED.exists():
        sys.exit(f"Missing {FEED} — the players feed is the id source.")
    feed = json.loads(FEED.read_text())
    clubs = {t["id"]: t["short_name"] for t in feed["teams"]}
    by_name = defaultdict(list)
    for e in feed["elements"]:
        by_name[e["web_name"].strip().lower()].append(e)
    return by_name, clubs


def resolve(name, position, club, by_name, clubs):
    """FPL element id for a drafted player, or None if unresolvable.

    Duplicate surnames are real (there are several Sánchezes), so where a name
    is ambiguous we narrow on the position and club recorded at the draft.
    """
    cands = by_name.get(str(name).strip().lower(), [])
    if not cands:
        return None
    if len(cands) == 1:
        return cands[0]["id"]
    narrowed = [
        c for c in cands
        if POS[c["element_type"]] == position and clubs.get(c["team"]) == club
    ]
    if len(narrowed) == 1:
        return narrowed[0]["id"]
    # Fall back to position alone — clubs change, positions rarely do.
    narrowed = [c for c in cands if POS[c["element_type"]] == position]
    return narrowed[0]["id"] if len(narrowed) == 1 else None


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    src = Path(sys.argv[1])
    if not src.exists():
        sys.exit(f"No such file: {src}")

    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl is required: pip install openpyxl")

    wb = openpyxl.load_workbook(src, data_only=True)
    if "All Sales" not in wb.sheetnames:
        sys.exit(f"Expected an 'All Sales' sheet; found {wb.sheetnames}")
    rows = [r for r in wb["All Sales"].iter_rows(values_only=True) if r and r[0]][1:]

    by_name, clubs = load_feed()
    identities = load_identities()

    unknown_owners = {r[3] for r in rows} - set(identities)
    if unknown_owners:
        sys.exit(f"No identity mapping for: {sorted(unknown_owners)}")

    teams = defaultdict(list)
    unresolved = []
    for name, position, club, owner, price in rows:
        pid = resolve(name, position, club, by_name, clubs)
        if pid is None:
            unresolved.append({"name": name, "position": position, "club": club,
                               "team": identities[owner]["team"]})
            continue
        teams[identities[owner]["key"]].append({
            "id": pid,
            "name": name,
            "position": position,
            "club": club,
            "price": float(price or 0),
        })

    by_key = {v["key"]: v["team"] for v in identities.values()}
    print(f"{len(rows)} drafted players across {len(teams)} teams")
    for key in sorted(teams):
        counts = defaultdict(int)
        for p in teams[key]:
            counts[p["position"]] += 1
        shape = " ".join(f"{k}{counts[k]}" for k in ("GK", "DEF", "MID", "FWD"))
        flag = "" if len(teams[key]) == 15 else f"  ⚠ {len(teams[key])} players"
        print(f"  {key:<4} {by_key[key]:<24} {shape}{flag}")

    if unresolved:
        print(f"\n⚠ {len(unresolved)} player(s) could not be resolved to an FPL id:")
        for u in unresolved:
            print(f"    {u}")
        print("  These would score nothing — fix before trusting any totals.")

    out = {
        "source": src.name,
        "teams": [
            {"key": key, "team": by_key[key],
             "squad": sorted(teams[key], key=lambda p: p["position"])}
            for key in sorted(teams)
        ],
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(out, indent=2))
    print(f"\nWrote {OUT.relative_to(ROOT.parent)}")
    return 1 if unresolved else 0


if __name__ == "__main__":
    sys.exit(main())
