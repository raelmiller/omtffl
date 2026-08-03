#!/usr/bin/env python3
"""Normalise OMTFFL draft history workbooks into one dataset.

The workbooks lay squads out as 5-column blocks (Position, Player, Club,
Cost, spacer) with the manager's name in the row above the header. Blocks
can sit anywhere on the sheet (side by side and/or stacked), so we scan
for every "Position/Player/Club/Cost" header and read downwards.

Outputs:
  data/draft_history.csv   - one row per sale
  data/draft_history.json  - same data plus per-season summaries
"""
import csv
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
RAW = ROOT / "data" / "raw"

FILES = {
    "2023-24": "OMTFFL_Draft_2324.xlsx",
    "2024-25": "OMTFFL_Draft_2425.xlsx",
    "2025-26": "OMTFFL_Draft_2526.xlsx",
}

VALID_POSITIONS = {"GK", "DEF", "MID", "FOR", "FWD"}

# same person, different name recorded between seasons (confirmed by Rael)
MANAGER_ALIASES = {"Ollie": "Oli", "Lucien": "Luke", "Georgeson": "Luke"}


def parse_cost(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    s = str(v).strip().replace("£", "")
    if not s:
        return None
    try:
        return round(float(s), 2)
    except ValueError:
        return None


def parse_sheet(ws, season):
    rows = []
    # find every header block anchor: cell == "Position" with "Player" to its right
    anchors = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if (
                str(ws.cell(r, c).value).strip().lower() == "position"
                and str(ws.cell(r, c + 1).value).strip().lower() == "player"
            ):
                anchors.append((r, c))

    for (hr, hc) in anchors:
        manager = None
        # manager name sits in the same column, 1-2 rows above the header
        for up in (1, 2):
            v = ws.cell(hr - up, hc).value
            if v is not None and str(v).strip():
                manager = str(v).strip()
                break
        if manager is None:
            manager = f"UNKNOWN(r{hr}c{hc})"
        manager = MANAGER_ALIASES.get(manager, manager)

        r = hr + 1
        while r <= ws.max_row:
            pos = ws.cell(r, hc).value
            pos = str(pos).strip().upper() if pos is not None else ""
            if pos not in VALID_POSITIONS:
                break  # end of block (blank row, "Amount remaining", next header)
            player = ws.cell(r, hc + 1).value
            club = ws.cell(r, hc + 2).value
            cost = parse_cost(ws.cell(r, hc + 3).value)
            player = str(player).strip() if player is not None else ""
            club = str(club).strip().upper() if club is not None else ""
            if player:
                rows.append(
                    {
                        "season": season,
                        "manager": manager,
                        "position": "FWD" if pos == "FOR" else pos,
                        "player": player,
                        "club": club,
                        "cost": cost if cost is not None else 0.0,
                    }
                )
            r += 1
    return rows


def main():
    all_rows = []
    for season, fname in FILES.items():
        wb = openpyxl.load_workbook(RAW / fname, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = parse_sheet(ws, season)
        all_rows.extend(rows)

    # summaries for validation
    summaries = {}
    for season in FILES:
        srows = [r for r in all_rows if r["season"] == season]
        managers = {}
        for r in srows:
            m = managers.setdefault(
                r["manager"], {"players": 0, "spent": 0.0, "by_pos": {}}
            )
            m["players"] += 1
            m["spent"] = round(m["spent"] + r["cost"], 2)
            m["by_pos"][r["position"]] = m["by_pos"].get(r["position"], 0) + 1
        summaries[season] = {
            "teams": len(managers),
            "sales": len(srows),
            "total_spent": round(sum(r["cost"] for r in srows), 2),
            "managers": managers,
        }

    out_csv = ROOT / "data" / "draft_history.csv"
    with out_csv.open("w", newline="") as f:
        w = csv.DictWriter(
            f, fieldnames=["season", "manager", "position", "player", "club", "cost"]
        )
        w.writeheader()
        w.writerows(all_rows)

    out_json = ROOT / "data" / "draft_history.json"
    out_json.write_text(json.dumps({"sales": all_rows, "summaries": summaries}, indent=1))

    print(json.dumps(summaries, indent=2))


if __name__ == "__main__":
    main()
